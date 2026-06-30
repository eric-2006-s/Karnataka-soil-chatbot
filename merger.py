import pandas as pd
import os
from openpyxl.styles import Font, PatternFill, Alignment

CSV_FILE = r"C:\Users\erics\OneDrive\Desktop\Export_Output.csv"
EXCEL_FILE = r"C:\Users\erics\OneDrive\Desktop\Export_Output.xlsx"

def csv_to_excel(csv_path, excel_path):
    df = pd.read_csv(csv_path)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")

        ws = writer.sheets["Sheet1"]

        # Bold green header row
        header_fill = PatternFill("solid", fgColor="2E7D32")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    print(f"Saved: {excel_path}")

if not os.path.exists(CSV_FILE):
    print(f"ERROR: File not found — {CSV_FILE}")
else:
    csv_to_excel(CSV_FILE, EXCEL_FILE)
    print("Done! Export_Output.xlsx saved to Desktop.")